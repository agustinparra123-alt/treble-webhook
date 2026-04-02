from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

FILE_NAME = "conversations.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["Conversation ID", "Phone", "Message", "Timestamp"])
    wb.save(FILE_NAME)

@app.route('/webhook', methods=['POST'])
def webhook():
    data = request.json
    print("Incoming:", data)

    conversation_id = data.get("conversation_id")
    phone = data.get("phone")
    message = data.get("message")
    timestamp = data.get("timestamp")

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([conversation_id, phone, message, timestamp])
    wb.save(FILE_NAME)

    return jsonify({"status": "ok"}), 200

@app.route('/')
def home():
    return "Webhook is live!"

if __name__ == "__main__":
    app.run()
